from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.db.models.signals import post_save
from django.views.generic.edit import CreateView, UpdateView
from django.contrib.auth.forms import UserCreationForm
from .forms import ProductForm, CustomUserCreationForm
from django.dispatch import receiver
from .models import Product
from django.contrib.auth.models import User
from .models import Category
from .serializers import ProductSerializer
from shahzodspace.models import Sale
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.decorators import api_view
from django.contrib.auth import login
from .models import UserProfile
from rest_framework.authtoken.models import Token
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from .forms import CategoryForm
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from decimal import Decimal
from .forms import SaleForm
from .forms import DateRangeForm
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.db.models import Sum, F, ExpressionWrapper, DecimalField
from django.db.models import Sum

def home(request):
    return render(request, 'shahzodspace/home.html')  # Путь к шаблону предположительно верный

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Проверяем, существует ли уже профиль для пользователя
            UserProfile.objects.get_or_create(user=user)  # Используем get_or_create, чтобы избежать дублирования
            login(request, user)
            return redirect('home')
    else:
        form = UserCreationForm()
    return render(request, 'shahzodspace/register.html', {'form': form})

@login_required
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            product = form.save(commit=False)
            product.user = request.user
            product.save()
            return redirect('product_list')
    else:
        form = ProductForm(user=request.user)
    return render(request, 'shahzodspace/product_form.html', {'form': form})




def product_list(request):
    products = Product.objects.filter(user=request.user)
    return render(request, 'shahzodspace/product_list.html', {'products': products})

class ProductCreate(LoginRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'shahzodspace/product_add.html'
    success_url = reverse_lazy('product_list')

    def form_valid(self, form):
        form.instance.user = self.request.user  # Установка пользователя для продукта
        return super().form_valid(form)

class ProductUpdate(UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'shahzodspace/product_form.html'  # Исправлен путь к шаблону
    success_url = reverse_lazy('product_list')

# Классы для REST API, предположительно правильные
class ProductList(generics.ListCreateAPIView):
    serializer_class = ProductSerializer

    def get_queryset(self):
        """Возвращает только товары текущего пользователя."""
        return Product.objects.filter(user=self.request.user)

class ProductDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer

@login_required
def profile(request):

    user = request.user
    user_profile, created = UserProfile.objects.get_or_create(user=user)  # Создаем UserProfile, если он не существует
    days_until_trial_ends = user_profile.days_until_trial_ends

    context = {
        'days_until_trial_ends': days_until_trial_ends
    }

      # Доход: сумма всех продаж
    total_income = Sale.objects.filter(product__user=request.user).aggregate(
        total=Sum(F('quantity') * F('sale_price')))['total'] or 0

    # Расход: сумма закупочной стоимости проданных товаров
    total_expense = Sale.objects.filter(product__user=request.user).annotate(
        cost=ExpressionWrapper(F('quantity') * F('product__purchase_price'), output_field=DecimalField())
    ).aggregate(total=Sum('cost'))['total'] or 0

    # Прибыль: разница между доходом и расходом
    total_profit = total_income - total_expense


    products_count = Product.objects.filter(user=request.user).count()
    registration_date = request.user.date_joined
    context = {
        'total_income': total_income,
        'total_expense': total_expense,
        'total_profit': total_profit,
        'products_count': products_count,
        'registration_date': registration_date,
        'user': request.user,
    }
    return render(request, 'shahzodspace/profile.html', context)

@login_required
def add_category(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.user = request.user  # Установка пользователя
            category.save()
            return redirect('category_list')  # Измените на URL, куда нужно перенаправить после создания
    else:
        form = CategoryForm()
    return render(request, 'shahzodspace/add_category.html', {'form': form})

@login_required
def category_list(request):
    # Получаем категории текущего пользователя
    categories = Category.objects.filter(user=request.user)
    # Отправляем эти категории в шаблон
    return render(request, 'shahzodspace/category_list.html', {'categories': categories})

def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk, user=request.user)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product, user=request.user)
        if form.is_valid():
            form.save()
            return redirect('product_list')
    else:
        form = ProductForm(instance=product, user=request.user)
    return render(request, 'shahzodspace/product_edit.html', {'form': form})

@login_required
@require_POST
def product_delete(request, pk):
            product = Product.objects.get(pk=pk)
            product.delete()
            return redirect('product_list')  # Вернитесь на страницу со списком товаров после удаления


@login_required
def product_list(request):
    search_query = request.GET.get('search', '')
    category_query = request.GET.get('category', None)
    model_query = request.GET.get('model', '')
    item_number_query = request.GET.get('item_number', '')  # Добавляем фильтр по номеру товара

    products = Product.objects.filter(user=request.user)

    if search_query:
        products = products.filter(name__icontains=search_query)

    if category_query:
        products = products.filter(category__id=category_query)

    if model_query:
        products = products.filter(model__icontains=model_query)  # Фильтрация по модели

    if item_number_query:
        products = products.filter(item_number__icontains=item_number_query)  # Фильтрация по номеру товара

    categories = Category.objects.filter(user=request.user)

    return render(request, 'shahzodspace/product_list.html', {
        'products': products,
        'categories': categories,
        'current_category': category_query,
        'current_model': model_query,
        'current_item_number': item_number_query  # Передаем текущий номер товара в шаблон
    })


@login_required
def export_products_to_excel(request):
    products = Product.objects.filter(user=request.user).values_list('name', 'category__name', 'model', 'quantity', 'purchase_price', 'selling_price')
    
    # Создание книги Excel и активного листа
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"

    columns = ['Наименование', 'Категория', 'Модель', 'Количество', 'Цена закупки', 'Цена продажи']
    row_num = 1

    # Заголовки столбцов
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Данные
    for product in products:
        row_num += 1
        for col_num, cell_value in enumerate(product, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
    
    # Установка ширины столбцов
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    # Создание HTTP-ответа
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'
    wb.save(response)

    return response

@login_required
def history_of_sales(request):
    form = DateRangeForm(request.GET or None, user=request.user)
    sales = Sale.objects.filter(product__user=request.user).order_by('-sale_date')
    
    if form.is_valid():
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
        category = form.cleaned_data['category']
        sales = sales.filter(sale_date__date__range=[start_date, end_date])
        
        if category:
            sales = sales.filter(product__category=category)
    
    # Расчеты
    total_income = sales.aggregate(Sum('sale_price'))['sale_price__sum'] or 0
    total_cost = sales.aggregate(Sum('product__purchase_price'))['product__purchase_price__sum'] or 0
    total_profit = total_income - total_cost

    return render(request, 'shahzodspace/history_of_sales.html', {
        'form': form,
        'sales': sales,
        'total_income': total_income,
        'total_cost': total_cost,
        'total_profit': total_profit,
    })


@login_required
def return_product(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)
    product = sale.product
    # Возврат количества товара
    product.quantity += sale.quantity
    product.save()
    # Удаление записи о продаже
    sale.delete()
    return redirect('history_of_sales')  # Вернуть пользователя на страницу истории продаж

@login_required
def sell_product(request, pk):
    product = get_object_or_404(Product, pk=pk, user=request.user)
    if request.method == 'POST':
        form = SaleForm(request.POST, product=product)
        if form.is_valid():
            sale = form.save(commit=False)
            sale.product = product
            # Расчет цены продажи с учетом скидки (если необходимо)
            product.quantity -= sale.quantity  # Уменьшаем количество товара на проданное
            product.save()
            sale.save()
            return redirect('history_of_sales')  # Направляем на страницу истории продаж
    else:
        form = SaleForm(product=product)
    return render(request, 'shahzodspace/sell_product.html', {'form': form, 'product': product})


@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Важно!
            messages.success(request, 'Ваш пароль был успешно обновлен!')
            return redirect('change_password')
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибку ниже.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'shahzodspace/change_password.html', {
        'form': form
    })

@login_required
def export_sales_report(request):
    # Получение данных о продажах текущего пользователя
    sales = Sale.objects.filter(product__user=request.user)

    # Создание книги Excel и активного листа
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "История продаж"

    # Заголовки для столбцов
    columns = ['Название товара', 'Количество', 'Цена продажи', 'Дата продажи']
    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Добавление данных в лист
    for row_num, sale in enumerate(sales, 2):
        ws.cell(row=row_num, column=1, value=sale.product.name)
        ws.cell(row=row_num, column=2, value=sale.quantity)
        ws.cell(row=row_num, column=3, value=sale.sale_price)
        ws.cell(row=row_num, column=4, value=sale.sale_date.strftime('%Y-%m-%d'))

    # Установка ширины столбцов
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    # Создание HTTP-ответа
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'
    wb.save(response)

    return response

def some_view(request):
    if not request.user.profile.is_trial_active and not request.user.profile.subscription_active:
        return redirect('subscription_required')  # Перенаправление на страницу с информацией о необходимости подписки
    # Остальная логика функции

def subscription_required_view(request):
    return render(request, 'subscription_required.html')